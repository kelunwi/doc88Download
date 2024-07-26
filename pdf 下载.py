import requests
import os
import glob
import fitz
import html
from urllib import parse
import json
import openpyxl
import shutil

class Base64Decoder:
    """自定义的Base64解码器类"""

    def __init__(self):
        # 初始化解码器所需的变量和字符映射
        self.m_base64Str = ''
        self.m_base64Count = 0
        self.m_END_OF_INPUT = -1
        self.m_base64Chars_r = [
            'P', 'J', 'K', 'L', 'M', 'N', 'O', 'I',
            # ... (其余字符省略)
        ]
        self.m_reverseBase64Chars = {element: index for index, element in enumerate(self.m_base64Chars_r)}

    def set_base64_str(self, s):
        """设置要解码的Base64字符串"""
        self.m_base64Str = s
        self.m_base64Count = 0

    def read_reverse_base64(self):
        """读取并返回下一个Base64字符的反向映射值"""
        if not self.m_base64Str:
            return -1
        while True:
            if self.m_base64Count >= len(self.m_base64Str):
                return -1
            next_character = self.m_base64Str[self.m_base64Count]
            self.m_base64Count += 1
            try:
                if self.m_reverseBase64Chars[next_character]:
                    return self.m_reverseBase64Chars[next_character]
            except KeyError:
                pass
            if next_character == 'P':
                return 0
        return -1

    @staticmethod
    def ntos(n):
        """将数字转换为对应的字符"""
        n = hex(n)[2:]
        n = "0" + n[-1] if len(n) == 1 else n
        return html.unescape('%' + n)

    def decode(self, s):
        """解码Base64字符串"""
        self.set_base64_str(s)
        result = ''
        done = False
        in_buffer = [0, 0, 0, 0]
        in_buffer[0] = self.read_reverse_base64()
        in_buffer[1] = self.read_reverse_base64()
        while (not done) and (in_buffer[0] != self.m_END_OF_INPUT) and (in_buffer[1] != self.m_END_OF_INPUT):
            in_buffer[2] = self.read_reverse_base64()
            in_buffer[3] = self.read_reverse_base64()
            result += self.ntos((((in_buffer[0] << 2) & 0xff) | in_buffer[1] >> 4))
            if in_buffer[2] != self.m_END_OF_INPUT:
                result += self.ntos((((in_buffer[1] << 4) & 0xff) | in_buffer[2] >> 2))
                if in_buffer[3] != self.m_END_OF_INPUT:
                    result += self.ntos((((in_buffer[2] << 6) & 0xff) | in_buffer[3]))
                else:
                    done = True
            else:
                done = True
            in_buffer[0] = self.read_reverse_base64()
            in_buffer[1] = self.read_reverse_base64()
        return parse.unquote(result)

class ExcelReader:
    """Excel文件读取类"""

    @staticmethod
    def read_column_from_excel(file_path, sheet_name, col_num):
        """从Excel文件中读取指定列的数据"""
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        column_data = []
        column_file_name_data = []
        for row in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=col_num).value
            cell_value2 = sheet.cell(row=row, column=2).value
            column_data.append(str(cell_value).replace('\n', ''))
            column_file_name_data.append(str(cell_value2).replace('\n', ''))
        return column_data, column_file_name_data

class PdfConverter:
    """PDF转换器类"""

    @staticmethod
    def pic2pdf(file_name):
        """将指定文件夹中的图片转换为PDF"""
        doc = fitz.open()
        for img in sorted(glob.glob(f"./{file_name}/*")):
            imgdoc = fitz.open(img)
            pdfbytes = imgdoc.convert_to_pdf()
            imgpdf = fitz.open("pdf", pdfbytes)
            doc.insert_pdf(imgpdf)
        if os.path.exists(f"./{file_name}/{file_name}.pdf"):
            os.remove(f"./{file_name}/{file_name}.pdf")
        doc.save(f"./{file_name}/{file_name}.pdf")
        doc.close()

class Doc88Downloader:
    """Doc88文档下载器类"""

    def __init__(self, excel_file_path, sheet_name, col_num):
        """初始化下载器，读取Excel数据"""
        self.excel_reader = ExcelReader()
        self.base64_decoder = Base64Decoder()
        self.pdf_converter = PdfConverter()
        self.p_code_data, self.file_name_data = self.excel_reader.read_column_from_excel(excel_file_path, sheet_name, col_num)

    def download_and_convert(self, start_index, step_size):
        """下载文档并转换为PDF"""
        header = {
            'Referer': 'http://m.doc88.com/',
            'User-Agent': 'Mozilla/5.0 (Linux; Android 9; ONEPLUS A6010 Build/PKQ1.180716.001; wv) AppleWebKit/537.36 (KHTML, '
                          'like Gecko) Version/4.0 Chrome/76.0.3809.89 Mobile Safari/537.36 T7/11.19 SP-engine/2.15.0 '
                          'baiduboxapp/11.19.0.11 (Baidu; P1 9) '
        }

        for i in range(start_index, len(self.p_code_data), step_size):
            i_end = min(i + step_size, len(self.p_code_data))
            for index, p_code in enumerate(self.p_code_data[i:i_end], start=i):
                print('下载中。。。')
                url = f'https://m.doc88.com/doc.php?act=info&p_code={p_code}&key=3854933de90d1dbb321d8ca29eac130a&v=1'
                result = requests.get(url, headers=header)
                return_data = result.text
                base64str = self.base64_decoder.decode(return_data)
                s = json.loads(base64str)

                if 'gif_host' not in s:
                    continue

                gif_host = s['gif_host']
                gif_urls = json.loads(s['gif_struct'])
                file_name = s['name'].replace('\u3000', ' ')
                file_name = ''.join(c for c in file_name if c not in r'\/:*?"<>|')

                # 创建新文件夹或清空已存在的文件夹
                if os.path.exists(file_name):
                    shutil.rmtree(file_name)
                os.mkdir(file_name)

                # 下载每个GIF文件
                for index, element in enumerate(gif_urls):
                    gif_url = f"{gif_host}/get-{element['u']}.gif"
                    result = requests.get(gif_url)
                    with open(f'./{file_name}/{str(index).zfill(7)}.gif', 'wb') as f:
                        f.write(result.content)

                # 将GIF转换为PDF
                self.pdf_converter.pic2pdf(file_name)
                print('下载完毕！')

if __name__ == "__main__":
    # 设置参数
    file_path = r'C:\Users\22725\Desktop\p_code_data - 副本.xlsx'
    sheet_name = 'Sheet'
    col_num = 1
    start_index = 50
    step_size = 10

    # 创建下载器实例并执行下载和转换
    downloader = Doc88Downloader(file_path, sheet_name, col_num)
    downloader.download_and_convert(start_index, step_size)